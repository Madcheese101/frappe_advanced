import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView
from frappe.utils.xlsxutils import handle_html
from frappe import _

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


# return xlsx file object
def make_xlsx(data, sheet_name, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=False)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = ILLEGAL_CHARACTERS_RE.sub("", value)

			clean_row.append(value)

		ws.append(clean_row)

	for col in ['C', 'D']:
		for cell in ws[col]:
			cell.number_format = '0.00'

	redFill = PatternFill(start_color='FFFF00',
               end_color='FFFF00',
               fill_type='solid')
	rows = ws.max_row	

	ws.conditional_formatting.add(f'A1:D{rows}',
             FormulaRule(formula=['AND($C1<>$D1,ISNUMBER($D1))'], stopIfTrue=False, fill=redFill))
	
	ws.auto_filter.ref = f'A1:D{rows}'

	# set the worksheet's sheet_view attribute to the new SheetView object
	ws.sheet_view.rightToLeft = True

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file